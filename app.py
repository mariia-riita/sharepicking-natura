import streamlit as st
import google.generativeai as genai
import pandas as pd
import io
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =========================================================================
# CONFIGURAÇÃO DA PÁGINA E DESIGN POPPINS
# =========================================================================
st.set_page_config(page_title="Cherry Picking - Natura", page_icon="💵", layout="wide")

st.markdown(
    """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;500;600;700&display=swap');

    html, body, [data-testid="stAppViewContainer"], .stApp {
        font-family: 'Poppins', sans-serif !important;
    }
    </style>
    """,
    unsafe_allow_html=True
)

st.title("💵 Agente de Sourcing: Cherry Picking")
st.write("Suba as planilhas oficiais ou CSVs dos fornecedores para gerar a consolidação automática.")

# =========================================================================
# SEGURANÇA: Configuração da Chave API oculta nos segredos do Streamlit
# =========================================================================
try:
    API_KEY = st.secrets["GEMINI_API_KEY"]
    genai.configure(api_key=API_KEY)
    chave_configurada = True
except Exception:
    chave_configurada = False

def obter_modelo_disponivel():
    """Identifica dinamicamente o nome exato do modelo ativo na sua conta para evitar erros 404."""
    try:
        modelos = [m.name for m in genai.list_models() if "generateContent" in m.supported_generation_methods]
        # Dá preferência para modelos flash/pro conhecidos
        for m in modelos:
            if "flash" in m or "pro" in m:
                return m
        if modelos:
            return modelos[0]
    except Exception:
        pass
    return "models/gemini-1.5-flash"

# Caixa de upload: Apenas planilhas Excel e CSV
arquivos_carregados = st.file_uploader(
    "Arraste as cotações aqui (Formatos aceitos: Excel .xlsx ou .csv):", 
    type=["xlsx", "csv"], 
    accept_multiple_files=True
)

def limpar_json_retornado(texto):
    texto = texto.strip()
    if texto.startswith("```"):
        linhas = texto.split("\n")
        if lines := [l for l in linhas if not l.startswith("```")]:
            texto = "\n".join(lines).strip()
    return texto

def limpar_valor(val):
    if pd.isna(val) or val is None:
        return None
    try:
        if isinstance(val, (int, float)):
            return float(val)
        texto = str(val).replace("R$", "").replace(" ", "").strip()
        if "." in texto and "," in texto:
            texto = texto.replace(".", "").replace(",", ".")
        elif "," in texto:
            texto = texto.replace(",", ".")
        return float(texto)
    except ValueError:
        return None

# Mapeamento e limpeza reserva via Python (descarta colunas de taxa/imposto)
def normalizar_colunas_backup(df, nome_fornecedor):
    col_map = {}
    colunas_diaria_candidatas = []
    
    for col in df.columns:
        c_lower = str(col).strip().lower()
        if any(k in c_lower for k in ["regi", "cd", "local"]):
            col_map[col] = "Região"
        elif any(k in c_lower for k in ["carg", "funç", "func"]):
            col_map[col] = "Cargo"
        elif any(k in c_lower for k in ["turn", "horar"]):
            col_map[col] = "Turnos"
        else:
            # Descarta colunas de taxa, imposto ou percentual
            if not any(bad in c_lower for bad in ["taxa", "imposto", "%", "percent"]):
                if any(good in c_lower for good in ["diaria", "diária", "preço", "preco", "valor"]):
                    colunas_diaria_candidatas.append(col)

    df_renamed = df.rename(columns=col_map)

    if "Região" not in df_renamed.columns:
        df_renamed["Região"] = "Geral"
    if "Cargo" not in df_renamed.columns:
        df_renamed["Cargo"] = "Ajudante Picking"
    if "Turnos" not in df_renamed.columns:
        df_renamed["Turnos"] = "1º Turno"

    # Seleciona a coluna de preço correta
    if colunas_diaria_candidatas:
        col_preco = colunas_diaria_candidatas[-1] # Prioriza a última coluna de preço (normalmente 'Valor da Diária')
        df_renamed[nome_fornecedor] = df_renamed[col_preco].apply(limpar_valor)
    else:
        outras = [c for c in df_renamed.columns if c not in ["Região", "Cargo", "Turnos"]]
        if outras:
            df_renamed[nome_fornecedor] = df_renamed[outras[-1]].apply(limpar_valor)

    return df_renamed[["Região", "Cargo", "Turnos", nome_fornecedor]]

# =========================================================================
# FUNÇÃO DE FORMATAÇÃO DO EXCEL
# =========================================================================
def estilizar_planilha_excel(df, fornecedores):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cherry Picking"
    
    colunas_completas = ["Região", "Cargo", "Turnos"] + fornecedores + ["Melhor Preço", "Fornecedor Vencedor"]
    ws.append(colunas_completas)
    
    idx_min_col = 4 + len(fornecedores)
    idx_winner_col = 5 + len(fornecedores)
    total_cols = idx_winner_col
    
    header_fill = PatternFill(start_color="9E472A", end_color="9E472A", fill_type="solid")
    header_font = Font(name="Poppins", size=11, bold=True, color="FFFFFF")
    border_cinza = Border(
        left=Side(style='thin', color='E0D8D3'), right=Side(style='thin', color='E0D8D3'),
        top=Side(style='thin', color='E0D8D3'), bottom=Side(style='thin', color='E0D8D3')
    )
    
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for col_idx in range(1, total_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border_cinza

    row_num = 2
    current_region = None
    color_index = 0
    fallback_colors = ["FDF2EE", "FAF4F0", "EBF2EE", "FDF5EC", "F5F2EB", "FCF9F2", "FDF2F4", "F4F5F6"]
    assigned_colors = {}

    for _, row in df.iterrows():
        regiao_original = row.get("Região", "")
        regiao = str(regiao_original).strip()
        
        if current_region is not None and regiao != current_region:
            row_num += 1
            
        current_region = regiao
        
        if regiao not in assigned_colors:
            assigned_colors[regiao] = fallback_colors[color_index % len(fallback_colors)]
            color_index += 1
            
        cor_fundo_regiao = assigned_colors[regiao]
        fill_linha = PatternFill(start_color=cor_fundo_regiao, end_color=cor_fundo_regiao, fill_type="solid")
        
        ws.cell(row=row_num, column=1, value=regiao_original)
        ws.cell(row=row_num, column=2, value=row.get("Cargo", ""))
        ws.cell(row=row_num, column=3, value=row.get("Turnos", ""))
        
        for col_idx, forn in enumerate(fornecedores, start=4):
            valor = limpar_valor(row.get(forn, None))
            ws.cell(row=row_num, column=col_idx, value=valor)
        
        col_let_forn_start = get_column_letter(4)
        col_let_forn_end = get_column_letter(3 + len(fornecedores))
        col_let_min = get_column_letter(idx_min_col)
        
        ws.cell(row=row_num, column=idx_min_col, value=f"=MIN({col_let_forn_start}{row_num}:{col_let_forn_end}{row_num})")
        ws.cell(row=row_num, column=idx_winner_col, value=f'=_xlfn.XLOOKUP({col_let_min}{row_num}, {col_let_forn_start}{row_num}:{col_let_forn_end}{row_num}, ${col_let_forn_start}$1:${col_let_forn_end}$1)')
        
        valores_linha = {col_idx: limpar_valor(row.get(forn, None)) for col_idx, forn in enumerate(fornecedores, start=4)}
        valores_validos = {col: val for col, val in valores_linha.items() if val is not None}
        coluna_vencedora = min(valores_validos, key=valores_validos.get) if valores_validos else None
        
        winner_fill = PatternFill(start_color="E6F0EA", end_color="E6F0EA", fill_type="solid")
        winner_font = Font(name="Poppins", size=10, bold=True, color="1E3D2F")
        normal_font = Font(name="Poppins", size=10, color="000000")
        
        for col_idx in range(1, total_cols + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.border = border_cinza
            cell.font = normal_font
            
            if col_idx == coluna_vencedora:
                cell.fill = winner_fill
                cell.font = winner_font
            else:
                cell.fill = fill_linha
                
            if col_idx in [1, 2, 3]:
                cell.alignment = left_align
            else:
                cell.alignment = center_align
                
            if 4 <= col_idx <= idx_min_col:
                cell.number_format = 'R$ #,##0.00'
                
        row_num += 1

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    ws.views.sheetView[0].showGridLines = True
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# =========================================================================
# OPERAÇÃO DA INTERFACE
# =========================================================================
if arquivos_carregados:
    fornecedores_detectados = []
    contexto_planilhas = ""
    dfs_backup = []

    for arquivo in arquivos_carregados:
        nome_fornecedor = arquivo.name.split(".")[0].replace("Cotações M.O.xlsx -", "").replace("Cotações M.O. -", "").strip()
        fornecedores_detectados.append(nome_fornecedor)
        
        try:
            if arquivo.name.endswith(".xlsx"):
                df_temp = pd.read_excel(arquivo)
            else:
                df_temp = pd.read_csv(arquivo)
            
            contexto_planilhas += f"\n--- PROPOSTA DO FORNECEDOR: {nome_fornecedor} ---\n"
            contexto_planilhas += df_temp.dropna(how="all").to_csv(index=False) + "\n"
            
            dfs_backup.append(normalizar_colunas_backup(df_temp, nome_fornecedor))
        except Exception as e:
            st.error(f"Erro ao ler {arquivo.name}: {e}")
        
    st.success(f"🤖 Agente: {len(arquivos_carregados)} fornecedores prontos: {', '.join(fornecedores_detectados)}")

    if st.button("🚀 Gerar Cherry Picking Mestre"):
        with st.spinner("🧠 Unificando planilhas e extraindo diárias oficiais..."):
            df_consolidado = None
            
            if chave_configurada:
                try:
                    nome_modelo = obter_modelo_disponivel()
                    config_segura_json = {"temperature": 0.1, "response_mime_type": "application/json"}
                    model = genai.GenerativeModel(model_name=nome_modelo)
                    
                    fornecedores_str = ", ".join([f'"{f}"' for f in fornecedores_detectados])
                    
                    prompt_unico = f"""
                    Você é um analista especialista em suprimentos da Natura.
                    Sua tarefa é consolidar os dados das planilhas de TODOS os fornecedores fornecidos abaixo em UMA ÚNICA TABELA CONSOLIDADA DE CHERRY PICKING.

                    Fornecedores a incluir como colunas de valores: {fornecedores_str}

                    Conteúdo das propostas:
                    {contexto_planilhas}

                    Regras Rígidas de Extração de Preços:
                    1. ATENÇÃO: Identifique o VALOR CHEIO DA DIÁRIA em R$ de cada fornecedor (Ex: 229.82, 355.36, 265.00).
                    2. NÃO confunda o valor da diária com percentuais de Taxa (Ex: 1.6 ou 0.0975) ou Imposto (Ex: 16.5). Ignore taxas e impostos.
                    3. Alinhe exatamente as mesmas linhas cruzando: Região, Cargo e Turnos.
                    4. Padronize os nomes de "Região", "Cargo" e "Turnos" de forma idêntica para todos os fornecedores (ex: "Murici / AL", "1º turno (Segunda à Sábado) - 06:00 - 14:00").
                    5. Se algum fornecedor não tiver cotação para uma linha, envie null.

                    Retorne APENAS uma lista de objetos JSON onde cada objeto representa uma linha com as chaves:
                    "Região", "Cargo", "Turnos", {fornecedores_str}
                    """
                    
                    resposta = model.generate_content(prompt_unico, generation_config=config_segura_json)
                    texto_json = limpar_json_retornado(resposta.text)
                    
                    dados_json = json.loads(texto_json)
                    df_consolidado = pd.DataFrame(dados_json)
                except Exception as e:
                    st.warning(f"⚠️ Processando via motor local (Erro na API IA: {e})")

            # Caso a API não esteja configurada ou falhe, executa o motor local tratado
            if df_consolidado is None and dfs_backup:
                df_consolidado = dfs_backup[0]
                for df_prox in dfs_backup[1:]:
                    df_consolidado = pd.merge(df_consolidado, df_prox, on=["Região", "Cargo", "Turnos"], how="outer")

            if df_consolidado is not None:
                if "Região" in df_consolidado.columns and "Turnos" in df_consolidado.columns:
                    df_consolidado = df_consolidado.sort_values(by=["Região", "Turnos"]).reset_index(drop=True)
                
                buffer_excel = estilizar_planilha_excel(df_consolidado, fornecedores_detectados)
                
                st.balloons()
                st.success("✨ Processo concluído com Sucesso!")
                
                st.write("📊 Prévia da Tabela Consolidada:")
                st.dataframe(df_consolidado, use_container_width=True)
                
                st.download_button(
                    label="📥 Clique aqui para baixar a Planilha Excel (.xlsx)",
                    data=buffer_excel,
                    file_name="Cherry_Picking_Consolidado_Final.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
